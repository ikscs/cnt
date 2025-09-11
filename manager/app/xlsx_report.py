import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

def mk_xlsx_report(title_text, data):
    df = pd.DataFrame(data)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    ws.insert_rows(1)

    # Merge cells for title
    num_columns = df.shape[1]
    merge_range = f"A1:{get_column_letter(num_columns)}1"
    ws.merge_cells(merge_range)

    # Set title text and format
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add Excel Table ===
    table_range = f"A2:{get_column_letter(num_columns)}{len(df)+2}"
    table = Table(displayName="MyTable", ref=table_range)

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Auto-fit column widths ===
    for col in ws.columns:
        max_length = 0
        col_idx = col[0].column
        col_letter = get_column_letter(col_idx)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_length + 4

    # Save final Excel to BytesIO
    final_output = BytesIO()
    wb.save(final_output)
#    final_output.seek(0)  # Important: reset pointer for reading

    return final_output.getbuffer()

if __name__ == '__main__':
    data = [
        {"Name": "Alice", "Age": 30, "City": "New York"},
        {"Name": "Bob", "Age": 25, "City": "San Francisco"},
        {"Name": "Charlie", "Age": 35, "City": "Chicago"},
    ]
    title_text = "Employee Report"

    result = mk_xlsx_report(title_text, data)

    with open(f'{title_text}.xlsx', 'wb') as f:
        f.write(result)

